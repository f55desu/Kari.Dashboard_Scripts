"""
wb_costs_api_downloader.py — Выгружает «Историю затрат» по рекламе Wildberries
через API и формирует xlsx, как файл из папки:
    ...\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_..._!!!\Затраты\Затраты ВБ\
    "История-затрат-Все-<FROM>T00_00_00+03_00-<TO>T00_00_00+03_00.xlsx"

Файл одностраничный, лист "История затрат", 7 колонок (заголовки в строке 1,
данные с 2-й строки, БЕЗ служебной строки "Период"):
    ID кампании | Кампания | Раздел | Дата списания |
    Источник списания | Сумма | Номер документа

Источник данных — Wildberries Advertising (Promotion) API:
    GET https://advert-api.wildberries.ru/adv/v1/upd?from=YYYY-MM-DD&to=YYYY-MM-DD
Эндпоинт возвращает историю списаний (расходов) по всем рекламным кампаниям.
Ограничение периода — не более 31 дня на запрос, поэтому период бьётся на чанки.

Авторизация — токен WB категории «Продвижение» (заголовок Authorization).
Берётся из окружения WB_ADVERT_TOKEN (или WB_API_TOKEN). Без токена — выход.

Запуск:
    python wb_costs_api_downloader.py [--output-dir PATH]
                                      [--date-from YYYY-MM-DD] [--date-to YYYY-MM-DD]
  --date-from / --date-to  границы периода. По умолчанию: to = сегодня,
                           from = сегодня − 45 дней (1,5 месяца до сегодня).
  --output-dir             куда сохранить файл (по умолчанию сетевая шара Taldykin).

⚠️  ПРИМЕРНЫЙ ВАРИАНТ (калибруется на реальном токене):
    Имена полей ответа /adv/v1/upd надо сверить вживую — особенно "Раздел"
    (в образце "Единая Ставка") и "Источник списания" (в образце "Баланс").
    Места стыка вынесены в FIELD_MAP / лейблы, фактические ключи пишутся в лог.
"""
import sys
import os
import json
import time
import logging
from datetime import date, timedelta, datetime
from pathlib import Path

import requests
from openpyxl import Workbook

if sys.stdout is not None:
    sys.stdout.reconfigure(encoding="utf-8")
if sys.stderr is not None:
    sys.stderr.reconfigure(encoding="utf-8")

PROJECT_DIR = Path(__file__).parent

# Ключи API берём из .env рядом со скриптом (python-dotenv). Если пакета нет —
# просто работаем с уже выставленными переменными окружения.
try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_DIR / ".env")
except Exception:
    pass

downloads_folder = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Taldykin"
DEFAULT_DOWNLOADS_DIR = Path(downloads_folder)

LOG_FILE = os.path.join(PROJECT_DIR, "wb_costs_api.log")
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8',
)
log = logging.getLogger(__name__)

# --- Wildberries Advertising API ---
WB_ADV_BASE = "https://advert-api.wildberries.ru"
UPD_URL = f"{WB_ADV_BASE}/adv/v1/upd"

# Токен WB категории «Продвижение». Сначала спец-переменная, затем общий токен.
WB_API_TOKEN = (os.environ.get("WB_ADVERT_TOKEN", "").strip()
                or os.environ.get("WB_API_TOKEN", "").strip())

CHUNK_DAYS = 31              # макс. период на один запрос /adv/v1/upd
REQUEST_PAUSE = 1.0
MAX_RETRIES = 6
RETRY_STATUS = {429, 500, 502, 503, 504}
DEFAULT_LOOKBACK_DAYS = 45   # окно по умолчанию: 1,5 месяца до сегодняшнего дня

SHEET_NAME = "История затрат"
COLUMNS = [
    "ID кампании",
    "Кампания",
    "Раздел",
    "Дата списания",
    "Источник списания",
    "Сумма",
    "Номер документа",
]

# Поле ответа /adv/v1/upd → колонка. ⚠️ калибруется на реальном ответе.
FIELD_MAP = {
    "ID кампании": ["advertId", "campaignId", "id"],
    "Кампания": ["campName", "name", "campaignName"],
    # "Раздел" в образце = "Единая Ставка". Прямого поля, скорее всего, нет —
    # пробуем явные алиасы, иначе подставляем тип кампании по ADVERT_TYPE_LABELS.
    "Раздел": ["section", "advertSubType", "subjectName"],
    "Дата списания": ["updTime", "date", "changeTime"],
    "Источник списания": ["paymentType", "type", "source"],
    "Сумма": ["updSum", "sum", "amount"],
    "Номер документа": ["updNum", "docNumber", "documentNumber"],
}

# Тип кампании (advertType) → подпись «Раздела» (в /adv/v1/upd нет поля «Раздел»).
# Проверено: у аккаунта кампании type=9 — это новый «Аукцион»/«Единая ставка».
ADVERT_TYPE_LABELS = {
    8: "Авто",
    9: "Единая Ставка",
}
RAZDEL_DEFAULT = "Единая Ставка"   # дефолт, если тип неизвестен


def _get(url: str, params: dict) -> object:
    """GET к WB Advertising API с ретраями на 429/5xx."""
    headers = {"Authorization": WB_API_TOKEN, "Content-Type": "application/json"}
    attempt = 0
    while True:
        attempt += 1
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=90)
        except requests.RequestException as e:
            if attempt >= MAX_RETRIES:
                raise
            wait = min(60, 5 * attempt)
            log.warning("Сетевая ошибка (%s); попытка %d/%d, пауза %ds", e, attempt, MAX_RETRIES, wait)
            time.sleep(wait)
            continue
        if resp.status_code in RETRY_STATUS:
            # Спец-случай: WB-шлюз отдаёт 503 "timeout" — у аккаунта слишком много
            # кампаний, /adv/v1/upd не успевает сформировать ответ за лимит шлюза
            # (~45с). Ретраи и уменьшение периода не помогают (даже 1 день падает).
            if resp.status_code == 503 and "timeout" in resp.text.lower():
                if attempt >= 3:
                    raise RuntimeError(
                        "WB /adv/v1/upd стабильно отвечает 503 timeout: эндпоинт не "
                        "успевает сформировать ответ из-за большого числа рекламных "
                        "кампаний на аккаунте. Это ограничение на стороне WB (не кода). "
                        "Варианты: выгружать этот файл браузерным экспортом, либо "
                        "обратиться в поддержку WB по таймауту /adv/v1/upd.")
                wait = 10
            elif attempt >= MAX_RETRIES:
                resp.raise_for_status()
            else:
                wait = 30 if resp.status_code == 429 else min(30, 3 * attempt)
            log.warning("HTTP %d (%s); попытка %d/%d, пауза %ds",
                        resp.status_code, resp.text[:120], attempt, MAX_RETRIES, wait)
            time.sleep(wait)
            continue
        resp.raise_for_status()
        time.sleep(REQUEST_PAUSE)
        # 204/пустой ответ → пустой список
        if not resp.content:
            return []
        return resp.json()


def _chunks(date_from: date, date_to: date, step: int):
    """Разбивает [from, to] на непересекающиеся окна ≤ step дней."""
    cur = date_from
    while cur <= date_to:
        end = min(date_to, cur + timedelta(days=step - 1))
        yield cur, end
        cur = end + timedelta(days=1)


def fetch_upd(date_from: date, date_to: date) -> list:
    """История списаний за период (по чанкам ≤31 дня). Возвращает список записей."""
    records = []
    logged = False
    for c_from, c_to in _chunks(date_from, date_to, CHUNK_DAYS):
        data = _get(UPD_URL, {"from": c_from.strftime("%Y-%m-%d"),
                              "to": c_to.strftime("%Y-%m-%d")})
        batch = data if isinstance(data, list) else (data.get("upd") or data.get("data") or [])
        if not logged and batch:
            log.info("Ключи записи /adv/v1/upd (для калибровки): %s", list(batch[0].keys()))
            logged = True
        records.extend(batch)
        log.info("/adv/v1/upd %s..%s: получено=%d, всего=%d",
                 c_from, c_to, len(batch), len(records))
    return records


def _pick(rec: dict, aliases: list):
    for a in aliases:
        if a in rec and rec[a] not in (None, ""):
            return rec[a]
    return ""


def _format_date(value):
    """ISO '2026-06-18T23:59:00+03:00' → '2026-06-18 23:59'. Иначе как есть."""
    if not value:
        return ""
    s = str(value)
    try:
        s_norm = s.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s_norm)
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return s


def _num(value):
    if value in (None, ""):
        return ""
    try:
        f = float(value)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return value


def build_row(rec: dict) -> dict:
    razdel = _pick(rec, FIELD_MAP["Раздел"])
    if razdel == "":
        razdel = ADVERT_TYPE_LABELS.get(rec.get("advertType"), RAZDEL_DEFAULT)
    return {
        "ID кампании": _num(_pick(rec, FIELD_MAP["ID кампании"])),
        "Кампания": _pick(rec, FIELD_MAP["Кампания"]),
        "Раздел": razdel,
        "Дата списания": _format_date(_pick(rec, FIELD_MAP["Дата списания"])),
        "Источник списания": _pick(rec, FIELD_MAP["Источник списания"]),
        "Сумма": _num(_pick(rec, FIELD_MAP["Сумма"])),
        "Номер документа": _num(_pick(rec, FIELD_MAP["Номер документа"])),
    }


def write_workbook(rows: list, dest: Path) -> None:
    """xlsx с листом 'История затрат': заголовки в строке 1, данные с 2-й."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for j, col in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=j, value=col)
    for i, rec in enumerate(rows, start=2):
        for j, col in enumerate(COLUMNS, start=1):
            val = rec.get(col, "")
            ws.cell(row=i, column=j, value=(None if val == "" else val))
    wb.save(dest)


def _filename(date_from: date, date_to: date) -> str:
    """Имя как у образца: 'История-затрат-Все-<FROM>T00_00_00+03_00-<TO>T00_00_00+03_00.xlsx'."""
    stamp = "T00_00_00+03_00"
    return (f"История-затрат-Все-{date_from.strftime('%Y-%m-%d')}{stamp}"
            f"-{date_to.strftime('%Y-%m-%d')}{stamp}.xlsx")


def download_wb_costs_report(date_from: date = None, date_to: date = None,
                             output_dir: Path = None) -> bool:
    log.info("Starting WB Costs (Advertising API) Downloader...")

    if not WB_API_TOKEN:
        msg = "WB_ADVERT_TOKEN / WB_API_TOKEN не задан — нужен токен WB «Продвижение». Добавьте в .env."
        log.error(msg)
        print(f"[ERROR] {msg}")
        return False

    if date_to is None:
        date_to = date.today()
    if date_from is None:
        date_from = date_to - timedelta(days=DEFAULT_LOOKBACK_DAYS)

    DOWNLOADS_DIR = Path(output_dir) if output_dir else DEFAULT_DOWNLOADS_DIR
    if output_dir:
        DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    else:
        try:
            DEFAULT_DOWNLOADS_DIR.stat()
        except Exception as e:
            print(f"[ERROR] Cannot access downloads folder: {downloads_folder}. {e}")
            return False

    log.info("Период: %s .. %s", date_from, date_to)

    try:
        records = fetch_upd(date_from, date_to)
        log.info("Записей истории затрат: %d", len(records))
        rows = [build_row(r) for r in records]
        dest = DOWNLOADS_DIR / _filename(date_from, date_to)
        write_workbook(rows, dest)
        log.info("SUCCESS: сохранено %d строк → %s", len(rows), dest)
        print(f"SUCCESS: saved {len(rows)} rows -> {dest}")
        return True

    except Exception as e:
        log.exception("Ошибка выгрузки истории затрат WB: %s", e)
        print(f"[ERROR] {e}")
        return False


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Download Wildberries ad cost history via Advertising API")
    parser.add_argument("--date-from", type=str, default=None,
                        help="Начало периода YYYY-MM-DD (по умолчанию сегодня − 45 дней, т.е. 1,5 месяца)")
    parser.add_argument("--date-to", type=str, default=None,
                        help="Конец периода YYYY-MM-DD (по умолчанию сегодня)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Куда сохранить файл (по умолчанию сетевая шара Taldykin)")
    args = parser.parse_args()

    d_from = datetime.strptime(args.date_from, "%Y-%m-%d").date() if args.date_from else None
    d_to = datetime.strptime(args.date_to, "%Y-%m-%d").date() if args.date_to else None

    success = download_wb_costs_report(date_from=d_from, date_to=d_to, output_dir=args.output_dir)
    sys.stdout.flush(); sys.stderr.flush()  # os._exit не сбрасывает буферы print()
    os._exit(0 if success else 1)
