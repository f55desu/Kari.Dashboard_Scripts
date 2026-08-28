import os
import re
import xlwings as xw
from datetime import datetime, timedelta

from pathlib import Path
import shutil
import win32com.client as win32
import time

import pandas as pd

import subprocess

# 🔧 Параметры (КОНСТАНТЫ ПУТЕЙ)
query_name = "Воронка"

folder_воронка = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\ВЫГРУЗКА воронка Озон"
folder_показатели_по_дням = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\Показатели по дням"
folder_затраты_Озон = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\Затраты\Озон. Затраты из Аналитики"
folder_затраты_Озон_NewFormat = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\Затраты\Озон. Затраты из Аналитики New Format"


def copy_to_directory(source_path, target_dir, filename):
    src = os.path.join(source_path, filename)
    dst = os.path.join(target_dir, filename)

    shutil.copy2(src, dst)
    print(f"Скопирован: {src} в {dst}")
    return dst


def get_latest_filename(folder, pattern=r'analytics_report_(\d{4}-\d{2}-\d{2})_\d{2}_\d{2}\.xlsx', format="%Y-%m-%d"):
    files = [f for f in os.listdir(folder) if re.match(pattern, f)]
    if not files:
        raise FileNotFoundError("Нет файлов, соответствующих шаблону")

    def extract_date(file):
        match = re.search(pattern, file)
        if match:
            return datetime.strptime(match.group(1), format)
        return datetime.min

    latest_file = max(files, key=extract_date)
    return latest_file


def get_latest_file_by_pattern(folder, pattern):
    files = [f for f in os.listdir(folder) if re.match(pattern, f)]
    if not files:
        raise FileNotFoundError(f"Нет файлов по шаблону: {pattern}")
    files.sort(key=lambda f: os.path.getmtime(os.path.join(folder, f)), reverse=True)
    return files[0]


def get_latest_n_files_by_pattern(folder, pattern, n=3):
    files = [f for f in os.listdir(folder) if re.match(pattern, f)]
    if len(files) < n:
        raise FileNotFoundError(f"Нужно минимум {n} файлов по шаблону: {pattern}, найдено: {len(files)}")
    files.sort(key=lambda f: os.path.getmtime(os.path.join(folder, f)), reverse=True)
    return files[:n]


def _apply_xlsx_inplace_edits(src):
    """Правит файл 'Аналитика продвижения_{Дата}.xlsx' in-place через openpyxl.

    Лист 'Statistics':
        - удалить столбец 'Название товара' (если есть)
        - переименовать 'Средняя стоимость клика, ₽' -> 'Стоимость клика, ₽'
        - заменить '-' на 0.0 в данных
    Лист 'Union':
        - переименовать 'Продажи в продвижении, ₽' -> 'Продажи, ₽'
        - переименовать 'Продано товаров, шт'      -> 'Заказы, шт'

    Заголовки находятся на 2-й строке (1-я — служебная "Период ..."), поэтому
    HEADER_ROW = 2 в нумерации openpyxl.
    """
    from openpyxl import load_workbook

    wb = load_workbook(src)
    HEADER_ROW = 2

    def _headers(ws):
        return [c.value for c in ws[HEADER_ROW]]

    # ----- Statistics -----
    if "Statistics" in wb.sheetnames:
        ws = wb["Statistics"]

        for col_to_drop in ("Название товара",):
            headers = _headers(ws)
            if col_to_drop in headers:
                idx = headers.index(col_to_drop) + 1  # 1-based
                ws.delete_cols(idx)
                print(f"  ✅ [Statistics] удалён столбец '{col_to_drop}'")
            else:
                print(f"  ⚠️ [Statistics] столбец '{col_to_drop}' не найден")

        # переименования
        rename_map = {
            "Средняя стоимость клика, ₽": "Стоимость клика, ₽",
        }
        for old, new in rename_map.items():
            headers = _headers(ws)
            if old in headers:
                idx = headers.index(old) + 1
                ws.cell(row=HEADER_ROW, column=idx, value=new)
                print(f"  ✅ [Statistics] '{old}' → '{new}'")
            else:
                print(f"  ⚠️ [Statistics] столбец '{old}' не найден")

        # замена '-' на 0.0 в данных
        replaced = 0
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=False):
            for cell in row:
                if cell.value == '-':
                    cell.value = 0.0
                    replaced += 1
        if replaced:
            print(f"  ✅ [Statistics] заменено '-' → 0.0 в {replaced} ячейках")
    else:
        print("  ⚠️ Лист 'Statistics' не найден в файле")

    # ----- Union -----
    if "Union" in wb.sheetnames:
        ws = wb["Union"]
        rename_map = {
            "Продажи в продвижении, ₽": "Продажи, ₽",
            "Продано товаров, шт":      "Заказы, шт",
        }
        for old, new in rename_map.items():
            headers = _headers(ws)
            if old in headers:
                idx = headers.index(old) + 1
                ws.cell(row=HEADER_ROW, column=idx, value=new)
                print(f"  ✅ [Union] '{old}' → '{new}'")
            else:
                print(f"  ⚠️ [Union] столбец '{old}' не найден")
    else:
        print("  ⚠️ Лист 'Union' не найден в файле")

    wb.save(src)
    print(f"💾 Правки сохранены в '{src}'")


def copy_and_convert_sku_statistics(downloads_path, filename):
    import os, re, shutil
    from datetime import datetime
    import pandas as pd

    src = os.path.join(downloads_path, filename)

    # Проверяем, есть ли уже CSV для этого файла (файл мог быть обработан предыдущим запуском)
    match_check = re.search(r"Аналитика продвижения_(\d{2}\.\d{2}\.\d{4})\.xlsx", filename)
    if match_check:
        date_obj_check = datetime.strptime(match_check.group(1), "%d.%m.%Y") - timedelta(days=1)
        csv_check = os.path.join(folder_затраты_Озон, date_obj_check.strftime("%d.%m.%Y") + ".csv")
        if os.path.exists(csv_check):
            print(f"⏭️ CSV уже существует: {csv_check} — пропускаем повторную обработку")
            return csv_check

    print(f"🔧 Применяем правки к xlsx in-place: {filename}")
    _apply_xlsx_inplace_edits(src)

    df = pd.read_excel(src, sheet_name="Statistics", header=1)
    column_name = "Инструмент"
    target_column_name = "Тип продвижения"
    if column_name in df.columns:
        df.rename(columns={column_name: target_column_name}, inplace=True)
        print(f"✅ Столбец '{column_name}' был изменен на '{target_column_name}'.")
    else:
        print(f"⚠️ Столбец '{column_name}' не найден.")

    # В новом формате Ozon (с 01.08.2026) колонки "Место размещения" нет — errors='ignore'
    df.drop(columns="Место размещения", inplace=True, errors='ignore')

    for col in ['CTR, %', 'Конверсия в корзину, %', 'Затраты на заказ, ₽', 'Стоимость клика, ₽']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Извлекаем дату из имени файла
    match = re.search(r"Аналитика продвижения_(\d{2}\.\d{2}\.\d{4})\.xlsx", filename)
    if not match:
        raise ValueError(f"Не удалось извлечь дату из имени файла {filename}")
    file_date = match.group(1)

    date_obj = datetime.strptime(file_date, "%d.%m.%Y") - timedelta(days=1)
    formatted_date = date_obj.strftime("%d.%m.%Y")

    csv_filename = f"{formatted_date}.csv"
    csv_path = os.path.join(downloads_path, csv_filename)

    num_cols = len(df.columns)
    semis = ";" * (num_cols - 1)

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        f.write(f"Период {formatted_date} - {formatted_date}{semis}\n")
        f.write(f"{semis}\n")
        df.to_csv(
            f,
            index=False,
            header=True,
            sep=';',
            decimal=',',
            float_format="%.2f",
            lineterminator="\r\n"
        )

    print(f"✅ CSV сохранён: {csv_filename}")

    dst_csv          = os.path.join(folder_затраты_Озон,            csv_filename)
    dst_xlsx_newfmt  = os.path.join(folder_затраты_Озон_NewFormat,  filename)

    shutil.copy2(csv_path, dst_csv)
    print(f"✅ CSV скопирован → 'Затраты из Аналитики': {csv_filename}")

    shutil.copy2(src, dst_xlsx_newfmt)
    print(f"✅ Изменённый xlsx скопирован → 'Затраты из Аналитики New Format': {filename}")

    return os.path.join(folder_затраты_Озон, csv_filename)


def duplicate_file_with_date_format(folder, filename):
    yesterday = datetime.now() - timedelta(days=1)
    new_filename = yesterday.strftime("%d.%m.%Y") + ".xlsx"
    src_path = os.path.join(folder, filename)
    dst_path = os.path.join(folder, new_filename)

    shutil.copy2(src_path, dst_path)
    print(f"✅ Создана копия файла: {new_filename}")
    return new_filename


def update_power_query_filter():
    latest_filename_воронка = get_latest_filename(folder_воронка)
    latest_filename_по_дням = get_latest_filename(folder_показатели_по_дням, pattern=r'(\d{2}.\d{2}.\d{4})', format="%d.%m.%Y")

    dated_filename_по_дням = duplicate_file_with_date_format(folder_показатели_по_дням, latest_filename_по_дням)
    dated_path = os.path.join(folder_показатели_по_дням, dated_filename_по_дням)

    print(f"📂 Актуальный файл воронки: {latest_filename_воронка}")
    print(f"📄 Работаем с файлом-копией: {dated_filename_по_дням}")

    app = None
    try:
        app = xw.App(visible=False)
        wb = app.books.open(dated_path)
        xl = wb.app.api
        xl_wb = xl.Workbooks(os.path.basename(dated_filename_по_дням))

        for q in xl_wb.Queries:
            if q.Name == query_name:
                m_code = q.Formula
                break
        else:
            raise ValueError("❌ Запрос не найден")

        updated_code = re.sub(
            r'Table\.SelectRows\([^)]*Name[^"]*"[^"]+\.xlsx"',
            lambda m: re.sub(r'"[^"]+\.xlsx"', f'"{latest_filename_воронка}"', m.group(0)),
            m_code
        )

        if updated_code == m_code:
            raise ValueError("⚠️ Не удалось заменить имя файла в фильтре")

        q.Formula = updated_code

        wb.save()
        wb.close()
        print("✏️ M-код обновлён и сохранён")

    finally:
        if app:
            app.quit()
        time.sleep(10)

    if not os.path.exists(dated_path):
        raise FileNotFoundError(f"❌ Файл не найден: {dated_path}")

    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False

    try:
        dated_path = os.path.abspath(dated_path)
        wb = excel.Workbooks.Open(dated_path)
        if wb is None:
            raise RuntimeError(f"❌ Excel не смог открыть файл: {dated_path}")

        time.sleep(5)
        wb.RefreshAll()
        print("✏️ Обновляем данные Power Querry")
        sheet = wb.Sheets("Лист3")
        for pivot in sheet.PivotTables():
            pivot.RefreshTable()
        excel.CalculateUntilAsyncQueriesDone()
        wb.Save()
        wb.Close()
        print(f"✅ Данные в '{dated_path}' обновлены")
    finally:
        excel.Quit()


def main():
    """Main execution function for Dashboard OZON Monday Wrapper (3 дня: пятница, суббота, воскресенье)"""
    logs = open('logs.log', 'a')
    logs.write(f'{datetime.now()} - OZON Dashboard Monday Wrapper ran\n')
    logs.close()

    copyFrom = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Taldykin"

    # --- Обрабатываем 3 файла "Аналитика продвижения" по очереди ---
    pattern_sku = r"Аналитика продвижения_(\d{2}\.\d{2}\.\d{4})\.xlsx"
    sku_files = get_latest_n_files_by_pattern(copyFrom, pattern_sku, n=3)
    for file_sku in sku_files:
        print(f"\n{'='*60}")
        print(f"📦 Обработка SKU-файла: {file_sku}")
        print(f"{'='*60}")
        try:
            copy_and_convert_sku_statistics(filename=file_sku, downloads_path=copyFrom)
        except Exception as e:
            print(f"❌ Ошибка при обработке '{file_sku}': {e}")
            print("   Файл пропущен, продолжаем со следующим...")

    # --- Копируем 3 файла воронки по очереди ---
    pattern_report = r"analytics_report_\d{4}-\d{2}-\d{2}_\d{2}_\d{2}\.xlsx"
    report_files = get_latest_n_files_by_pattern(copyFrom, pattern_report, n=3)
    for file_report in reversed(report_files):
        copy_to_directory(copyFrom, folder_воронка, file_report)

    # --- ОПЦИОНАЛЬНО: загрузка листа "Union" свежего файла в PostgreSQL ---
    try:
        from ozon_union_db import try_upload_latest_to_postgres
        try_upload_latest_to_postgres()
    except Exception as e:
        print(f"⚠️ Шаг загрузки в PostgreSQL пропущен (не критично): {e}")
        try:
            with open('logs.log', 'a') as _l:
                _l.write(f'{datetime.now()} - PostgreSQL upload skipped: {e}\n')
        except Exception:
            pass

    # update_power_query_filter()

    logs = open('logs.log', 'a')
    logs.write(f'{datetime.now()} - OZON Dashboard Monday Wrapper completed\n')
    logs.close()


# ▶️ Запуск всех этапов
if __name__ == "__main__":
    main()
