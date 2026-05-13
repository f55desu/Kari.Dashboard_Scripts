import os
import re
import shutil
import zipfile
from datetime import datetime
import xlwings as xw

from openpyxl import load_workbook

import time
import win32com.client as win32

import pandas as pd


# 🔧 Пути
downloads_folder = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Taldykin"
current_folder = os.getcwd()

folder_wb = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\ВЫГРУЗКА воронка ВБ"
folder_zatraty_wb = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\Затраты\Затраты ВБ"
folder_week = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\Показатели по неделям ВБ"
folder_campaign_info = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!"

# folder_wb = "ВЫГРУЗКА воронка ВБ"
# folder_zatraty_wb = "Затраты\\Затраты ВБ"
# folder_week = "Показатели по неделям ВБ"
# folder_campaign_info = current_folder

file_voronka = ''
query_name = "ВЫГРУЗКА воронка ВБ"

# 📦 Этап 1: Распаковать архивы и переименовать файлы
def extract_latest_3_wb_files():
    # Находим все zip-файлы, соответствующие паттерну
    zips = [f for f in os.listdir(downloads_folder) if f.endswith('.zip') and re.match(r"\d{2}\.\d{2} с \d{2}\.\d{2}\.\d{4} по \d{2}\.\d{2}\.\d{4}\.zip", f)]
    if not zips:
        raise FileNotFoundError("Нет zip-файлов нужного формата в Загрузках")

    new_files = []  # Список для хранения путей к перемещённым файлам

    # Перебираем последние 3 архива
    for i, zip_file in enumerate(sorted(zips, key=lambda x: os.path.getmtime(os.path.join(downloads_folder, x)), reverse=True)[:3]):
        zip_path = os.path.join(downloads_folder, zip_file)
        
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # Извлекаем список файлов из архива
            extracted_files = zip_ref.namelist()
            # Ищем только .xlsx файлы
            xlsx_files = [f for f in extracted_files if f.endswith('.xlsx')]
            
            if not xlsx_files:
                raise FileNotFoundError(f"Нет .xlsx файлов в архиве: {zip_file}")
            
            # Берём первый (единственный) файл .xlsx из архива
            latest_xlsx = xlsx_files[0]
            
            # Извлекаем файл
            zip_ref.extract(latest_xlsx, downloads_folder)
            
            # Переименование файла
            src = os.path.join(downloads_folder, latest_xlsx)
            
            # Преобразование имени: 01.02 с 01.02.2026 по 01.02.2026.zip -> 01.02 с 1-2-2026 по 1-2-2026.xlsx
            # Извлекаем части из имени zip-файла
            match = re.search(r"(\d{2}\.\d{2}) с (\d{2})\.(\d{2})\.(\d{4}) по (\d{2})\.(\d{2})\.(\d{4})\.zip", zip_file)
            if not match:
                raise ValueError(f"Невозможно извлечь даты из имени zip-файла: {zip_file}")
            
            prefix = match.group(1)  # 01.02
            day1, month1, year1 = int(match.group(2)), int(match.group(3)), int(match.group(4))  # убираем ведущие нули через int()
            day2, month2, year2 = int(match.group(5)), int(match.group(6)), int(match.group(7))
            
            # Формируем новое имя файла
            new_name = f"{prefix} с {day1}-{month1}-{year1} по {day2}-{month2}-{year2}.xlsx"
            dst = os.path.join(folder_wb, new_name)  # Путь назначения
            
            shutil.move(src, dst)  # Перемещаем файл
            new_files.append(dst)  # Добавляем путь в список
            
            print(f"✅ Этап {i+1}: файл {new_name} перемещён в ВЫГРУЗКА воронка ВБ")
    new_files.reverse() # Переворачиваем список в обратную сторону от старшего к младшему файлу
    # Возвращаем пути к трем последним файлам
    return new_files[0], new_files[1], new_files[2]

def unify_columns(filepath, ethalon=os.path.join(downloads_folder, 'Воронка ВБ Эталон.xlsx'), sheet_name='Товары', header_row=2):
    """Приводит файл к структуре эталона: удаляет лишние столбцы И добавляет
       отсутствующие из эталона как пустые (см. 07.05.2026 — ВБ убрал «Остатки склад ВБ, шт»
       и «Остатки МП, шт», из-за чего Power Query терял данные за свежие дни)."""
    ethalon_df = pd.read_excel(ethalon, sheet_name=sheet_name, header=1)
    ethalon_headers = ethalon_df.columns.tolist()

    wb = load_workbook(filename=filepath)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[header_row]]

    # 1) Удаляем столбцы, которых нет в эталоне
    columns_to_delete = [h for h in headers if h not in ethalon_headers]
    for col in reversed(columns_to_delete):
        col_index = headers.index(col) + 1
        ws.delete_cols(col_index)
    headers = [cell.value for cell in ws[header_row]]

    # 2) Добавляем отсутствующие столбцы из эталона (на нужные позиции)
    for target_idx, eth_col in enumerate(ethalon_headers, start=1):
        if eth_col not in headers:
            ws.insert_cols(target_idx)
            ws.cell(row=header_row, column=target_idx, value=eth_col)
            headers.insert(target_idx - 1, eth_col)
            print(f"  ➕ Добавлен отсутствующий столбец '{eth_col}' (позиция {target_idx})")

    # 3) "." → "," в числовых строках
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            val = cell.value
            if isinstance(val, str):
                try:
                    float(val.replace(',', '.'))
                    if '.' in val:
                        cell.value = val.replace('.', ',')
                except ValueError:
                    pass

    wb.save(filepath)
    print(f"✅ Столбцы унифицированы (удалено лишних: {len(columns_to_delete)}): {filepath}")

def delete_columns_by_header(filepath, sheet_name="Товары", target_headers=None, header_row=2):
    if target_headers is None:
        target_headers = [
            "Добавили в отложенные",
            "Добавили в отложенные (предыдущий период)",
            "Заказали ВБ клуб, шт",
            "Заказали ВБ клуб, шт (предыдущий период)",
            "Выкупили ВБ клуб, шт",
            "Выкупы ВБ клуб, шт (предыдущий период)",
            "Отменили ВБ клуб, шт",
            "Отменили ВБ клуб, шт (предыдущий период)",
            "Процент выкупа ВБ клуб",
            "Процент выкупа ВБ клуб (предыдущий период)",
            "Заказали на сумму ВБ клуб, ₽",
            "Заказали на сумму ВБ клуб, ₽ (предыдущий период)",
            "Динамика суммы заказов ВБ клуб, ₽",
            "Выкупили на сумму ВБ клуб, ₽",
            "Выкупили на сумму ВБ клуб, ₽ (предыдущий период)",
            "Отменили на сумму ВБ клуб, ₽",
            "Отменили на сумму ВБ клуб, ₽ (предыдущий период)",
            "Средняя цена ВБ клуб, ₽",
            "Средняя цена ВБ клуб, ₽ (предыдущий период)",
            "Среднее количество заказов в день ВБ клуб, шт",
            "Среднее количество заказов в день ВБ клуб, шт (предыдущий период)"
        ]

    wb = load_workbook(filename=filepath)
    ws = wb[sheet_name]

    # Получаем заголовки из нужной строки (по умолчанию вторая)
    headers = [cell.value for cell in ws[header_row]]

    # Сохраняем индексы столбцов, которые нужно удалить
    col_indexes_to_delete = [
        idx + 1 for idx, value in enumerate(headers)
        if value in target_headers
    ]

    # Удаляем столбцы справа налево (иначе индексы сместятся)
    for col_idx in sorted(col_indexes_to_delete, reverse=True):
        ws.delete_cols(col_idx)

    wb.save(filepath)
    print(f"✅ Столбцы удалены: {', '.join(target_headers)}")

# 💰 Этап 2: Копировать файл История-затрат
def copy_latest_cost_history():
    # Новый паттерн для поиска файла "Не определено"
    pattern = r"История-затрат-Не определено-.*\.xlsx"
    
    # Ищем файлы, соответствующие паттерну
    files = [f for f in os.listdir(downloads_folder) if re.match(pattern, f)]
    
    # Если файлы не найдены, выбрасываем ошибку
    if not files:
        raise FileNotFoundError("Файл История-затрат не найден")
    
    # Находим последний файл по времени изменения
    latest = max(files, key=lambda f: os.path.getmtime(os.path.join(downloads_folder, f)))

    # Путь к последнему найденному файлу
    latest_file_path = os.path.join(downloads_folder, latest)

    # Переименование файла в новый паттерн "Все"
    renamed_file = re.sub(r"История-затрат-Не определено", "История-затрат-Все", latest)
    renamed_file_path = os.path.join(downloads_folder, renamed_file)

    # Переименовываем файл
    os.rename(latest_file_path, renamed_file_path)
    print(f"✅ Файл переименован в: {renamed_file}")

    # Очистка целевой папки
    for f in os.listdir(folder_zatraty_wb):
        os.remove(os.path.join(folder_zatraty_wb, f))

    # Копирование переименованного файла в целевую папку
    shutil.copy2(renamed_file_path, folder_zatraty_wb)
    
    print(f"✅ Этап 2: файл {renamed_file} скопирован в Затраты ВБ и папка очищена")

# 📈 Этап 3: Найти файл с максимальной неделей и обновить
def update_latest_week_file():
    pattern = r"Показы и затраты_(\d+)\sнеделя\.xlsx"
    week_files = [f for f in os.listdir(folder_week) if re.match(pattern, f)]
    if not week_files:
        raise FileNotFoundError("Файлы 'Показы и затраты' не найдены")

    def extract_week(f): return int(re.search(pattern, f).group(1))
    latest_week_file = max(week_files, key=extract_week)

    full_path = os.path.abspath(os.path.join(folder_week, latest_week_file))
    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Файл не найден: {full_path}")
    
    if file_voronka.startswith('01'):
        wb = xw.Book(full_path)
        xl = wb.app.api
        xl_wb = xl.Workbooks(os.path.basename(full_path))

        # Найти запрос
        for q in xl_wb.Queries:
            if q.Name == query_name:
                m_code = q.Formula
                break
        else:
            raise ValueError(f"Запрос '{query_name}' не найден")

        # Шаги 3.2-3.3: Заменить имя в шаге "Переименованные столбцы"
        # 🔽 Очистка заголовка из A1
        # yesterday = datetime.now() + timedelta(weeks=1) - timedelta(days=1) # TEST
        # yesterday = datetime.now() - timedelta(days=1) # RIGHT VERSION
        # day = yesterday.day
        # month = yesterday.month
        # year = yesterday.year
        new_filename = file_voronka

        # 🧠 Найти текущее имя файла в M-коде
        old_filename_match = re.search(
            r'"(\d{2}\.\d{2} с \d{1,2}-\d{1,2}-\d{4} по \d{1,2}-\d{1,2}-\d{4}\.xlsx)"',
            m_code
        )

        if not old_filename_match:
            raise ValueError("Не найдено старое имя файла в шаге 'Переименованные столбцы'")

        old_filename = old_filename_match.group(1)

        # 🔁 Безопасная замена
        m_code = m_code.replace(f'"{old_filename}"', f'"{new_filename}"')
        print(f"🔁 Заменено имя файла: {old_filename} → {new_filename}")

    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False  # Отключает предупреждения Excel
    try:
        wb = excel.Workbooks.Open(full_path)
        wb.RefreshAll()
        wb.Save()
        wb.Close()
        print(f"✅ Этап 3: файл '{latest_week_file}' обновлён")
    finally:
        excel.Quit()

# 📤 Этап 4: Переместить статусы_отчет.xlsx → Campaign_Info.xlsx
def move_campaign_info():
    src_file = os.path.join(downloads_folder, "статусы_отчет.xlsx")
    dst_file = os.path.join(folder_campaign_info, "Campaign_Info.xlsx")
    if not os.path.exists(src_file):
        raise FileNotFoundError("Файл статусы_отчет.xlsx не найден")
    shutil.move(src_file, dst_file)
    print("✅ Этап 4: Campaign_Info.xlsx перемещён в рабочую папку")

# 📊 Этап 5: Обновить Затраты ВБ_2.xlsx
def update_zatraty_wb2():
    file = os.path.join(folder_campaign_info, "Затраты ВБ_2.xlsx")
    if not os.path.exists(file):
        raise FileNotFoundError("Файл Затраты ВБ_2.xlsx не найден в рабочей папке")

    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False  # Отключает предупреждения Excel
    try:
        wb = excel.Workbooks.Open(file)
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        time.sleep(5)

        sheet = wb.Sheets["Лист1"]
        for pivot in sheet.PivotTables():
            pivot.RefreshTable()
        
        time.sleep(5)  # ⏳ Подождать 5 секунд
        # excel.CalculateUntilAsyncQueriesDone()  # Дожидаемся завершения обновления
        wb.Save()
        print("✅ Этап 6: Затраты ВБ_2.xlsx обновлён")
    finally:
        wb.Close()
        excel.Quit()


def main():
    """Main execution function for Dashboard WB Wrapper (Monday version)"""
    logs = open('logs.log', 'a')
    logs.write(f'{datetime.now()} - WB Dashboard Wrapper ran\n')

    voronka_wb_1, voronka_wb_2, voronka_wb_3 = extract_latest_3_wb_files() # 1
    unify_columns(voronka_wb_1)
    unify_columns(voronka_wb_2)
    unify_columns(voronka_wb_3)
    # copy_latest_cost_history() # 2
    update_latest_week_file()
    update_zatraty_wb2()

    # Запускаем скрипт
    # subprocess.run(['DB_WB_1.1.exe'], check=True)

    logs = open('logs.log', 'a')
    logs.write(f'{datetime.now()} - WB Dashboard Wrapper completed\n')
    logs.close()

# ▶️ Запуск всех этапов
if __name__ == "__main__":
    main()

