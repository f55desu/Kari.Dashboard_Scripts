"""
wb_funnel_api_downloader.py — Выгружает «Воронку продаж» Wildberries через API
и формирует xlsx с листом "Товары" (детальный отчёт по карточкам товаров),
по аналогии с браузерной выгрузкой из папки:
    ...\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_..._!!!\ВЫГРУЗКА воронка ВБ\
    "DD.MM с D-M-YYYY по D-M-YYYY.xlsx"

Лист "Товары" — 46 колонок: строка 1 = "Детальный отчет воронки продаж по
карточкам товаров", заголовки во 2-й строке, данные с 3-й.

⚠️  ВАЖНО: WB вывел из строя прежний синхронный эндпоинт
    /api/v2/nm-report/detail (теперь 404). Воронка стала АСИНХРОННЫМ
    CSV-отчётом:
      POST /api/v2/nm-report/downloads            — создать отчёт (reportType
                                                    = DETAIL_HISTORY_REPORT)
      GET  /api/v2/nm-report/downloads            — список со статусами
      GET  /api/v2/nm-report/downloads/file/{id}  — ZIP с CSV
    Новый CSV содержит по nmID за день только базовые метрики воронки:
      nmID, dt, openCardCount, addToCartCount, ordersCount, ordersSumRub,
      buyoutsCount, buyoutsSumRub, cancelCount, cancelSumRub,
      addToCartConversion, cartToOrderConversion, buyoutPercent,
      addToWishlist, currency
    Поэтому:
      • Артикул продавца / Название / Предмет / Бренд — дотягиваем из Content
        API (content/v2/get/cards/list) по nmID.
      • «Предыдущий период» — второй отчёт за (date-1) и join по nmID.
      • Показы, CTR, рейтинги, доля в выручке, средняя цена, остатки, время
        доставки, локальные заказы — В API ВОРОНКИ НЕТ → колонки оставлены
        пустыми (источник в WB UI другой). Можно добить отдельными эндпоинтами.

Авторизация — токен WB (заголовок Authorization), категории «Аналитика» и
«Контент». Берётся из окружения WB_API_TOKEN (.env).

Запуск:
    python wb_funnel_api_downloader.py [--output-dir PATH] [--date YYYY-MM-DD]
  --date        дата данных (по умолчанию сегодня; период = [date, date]).
  --output-dir  куда сохранить файл (по умолчанию сетевая шара Taldykin).
"""
import sys
import os
import io
import csv
import json
import time
import uuid
import zipfile
import logging
from datetime import date, timedelta, datetime
from pathlib import Path

import requests
from openpyxl import Workbook

if sys.stdout is not None:
    sys.stdout.reconfigure(encoding="utf-8")
if sys.stderr is not None:
    sys.stderr.reconfigure(encoding="utf-8")

PROJECT_DIR = Path(__file__).parent

# Ключи из .env рядом со скриптом (если есть python-dotenv).
try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_DIR / ".env")
except Exception:
    pass

downloads_folder = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Taldykin"
DEFAULT_DOWNLOADS_DIR = Path(downloads_folder)

LOG_FILE = os.path.join(PROJECT_DIR, "wb_funnel_api.log")
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8',
)
log = logging.getLogger(__name__)

# --- Wildberries API ---
WB_ANALYTICS_BASE = "https://seller-analytics-api.wildberries.ru"
DOWNLOADS_URL = f"{WB_ANALYTICS_BASE}/api/v2/nm-report/downloads"
DOWNLOADS_FILE_URL = f"{WB_ANALYTICS_BASE}/api/v2/nm-report/downloads/file/{{id}}"
CONTENT_CARDS_URL = "https://content-api.wildberries.ru/content/v2/get/cards/list"
# Остатки (асинхронный отчёт): create → status → download (JSON)
WAREHOUSE_REMAINS_URL = f"{WB_ANALYTICS_BASE}/api/v1/warehouse_remains"
WAREHOUSE_STATUS_URL = f"{WB_ANALYTICS_BASE}/api/v1/warehouse_remains/tasks/{{id}}/status"
WAREHOUSE_DOWNLOAD_URL = f"{WB_ANALYTICS_BASE}/api/v1/warehouse_remains/tasks/{{id}}/download"
# Псевдо-склады в отчёте остатков
WH_TOTAL_NAME = "Всего находится на складах"   # → Остатки «Склад WB»
WH_MP_NAME = "Маркетплейс"                       # → Остатки МП (FBS)

WB_API_TOKEN = os.environ.get("WB_API_TOKEN", "").strip()

MAX_RETRIES = 6
RETRY_STATUS = {429, 500, 502, 503, 504}
POLL_INTERVAL = 8            # сек между опросами статуса отчёта
POLL_TIMEOUT = 360          # сек максимум ожидания готовности отчёта
TIMEZONE = "Europe/Moscow"
AGG_LEVEL = "day"
CONTENT_PAGE = 100

SHEET_TITLE = "Детальный отчет воронки продаж по карточкам товаров"

# Заголовки листа "Товары" — 46 колонок, точный порядок 1:1 с оригиналом.
TOVARY_COLUMNS = [
    "Артикул продавца",
    "Артикул WB",
    "Название",
    "Предмет",
    "Бренд",
    "Удаленный товар",
    "Рейтинг карточки",
    "Рейтинг по отзывам",
    "Показы",
    "Показы (предыдущий период)",
    "Доля карточки в выручке",
    "Доля карточки в выручке (предыдущий период)",
    "Переходы в карточку",
    "Переходы в карточку (предыдущий период)",
    "Положили в корзину",
    "Положили в корзину (предыдущий период)",
    "Заказали товаров, шт",
    "Заказали товаров, шт (предыдущий период)",
    "Выкупили, шт",
    "Выкупы, шт (предыдущий период)",
    "Отменили, шт",
    "Отменили, шт (предыдущий период)",
    "Конверсия в корзину, %",
    "Конверсия в корзину, % (предыдущий период)",
    "Конверсия в заказ, %",
    "Конверсия в заказ, % (предыдущий период)",
    "Процент выкупа",
    "Процент выкупа (предыдущий период)",
    "Заказали на сумму, ₽",
    "Заказали на сумму, ₽ (предыдущий период)",
    "Динамика суммы заказов, ₽",
    "Выкупили на сумму, ₽",
    "Выкупили на сумму, ₽ (предыдущий период)",
    "Отменили на сумму, ₽",
    "Отменили на сумму, ₽ (предыдущий период)",
    "Средняя цена, ₽",
    "Средняя цена, ₽ (предыдущий период)",
    "Среднее количество заказов в день, шт",
    "Среднее количество заказов в день, шт (предыдущий период)",
    "Остатки «Склад WB», шт",
    "Остатки МП, шт",
    "Сумма остатков на складах, ₽",
    "Среднее время доставки",
    "Среднее время доставки (предыдущий период)",
    "Локальные заказы, %",
    "Локальные заказы, % (предыдущий период)",
]


def _request(method: str, url: str, *, json_body: dict = None, stream: bool = False) -> requests.Response:
    """HTTP к WB API с ретраями на 429/5xx."""
    headers = {"Authorization": WB_API_TOKEN, "Content-Type": "application/json"}
    attempt = 0
    while True:
        attempt += 1
        try:
            resp = requests.request(method, url, headers=headers,
                                    data=json.dumps(json_body) if json_body is not None else None,
                                    timeout=120, stream=stream)
        except requests.RequestException as e:
            if attempt >= MAX_RETRIES:
                raise
            wait = min(60, 5 * attempt)
            log.warning("Сетевая ошибка (%s); попытка %d/%d, пауза %ds", e, attempt, MAX_RETRIES, wait)
            time.sleep(wait)
            continue
        if resp.status_code in RETRY_STATUS:
            if attempt >= MAX_RETRIES:
                resp.raise_for_status()
            wait = 30 if resp.status_code == 429 else min(30, 3 * attempt)
            log.warning("HTTP %d (%s); попытка %d/%d, пауза %ds",
                        resp.status_code, resp.text[:120], attempt, MAX_RETRIES, wait)
            time.sleep(wait)
            continue
        resp.raise_for_status()
        return resp


# ---------------------------------------------------------------------------
# Content API: nmID → артикул/название/предмет/бренд
# ---------------------------------------------------------------------------
def fetch_content_map() -> dict:
    """Строит {nmID(str): {vendorCode, title, subjectName, brand}} по всем карточкам."""
    out = {}
    cursor = {"limit": CONTENT_PAGE}
    for _ in range(1000):  # предохранитель
        body = {"settings": {"cursor": cursor, "filter": {"withPhoto": -1}}}
        try:
            data = _request("POST", CONTENT_CARDS_URL, json_body=body).json()
        except Exception as e:
            log.warning("Content API недоступен (%s) — артикул/название/бренд будут пустыми", e)
            return out
        cards = data.get("cards") or []
        for c in cards:
            out[str(c.get("nmID"))] = {
                "vendorCode": c.get("vendorCode", ""),
                "title": c.get("title", ""),
                "subjectName": c.get("subjectName", ""),
                "brand": c.get("brand", ""),
            }
        cur = data.get("cursor") or {}
        if len(cards) < CONTENT_PAGE:
            break
        cursor = {"limit": CONTENT_PAGE, "updatedAt": cur.get("updatedAt"), "nmID": cur.get("nmID")}
    log.info("Content: карточек в справочнике: %d", len(out))
    return out


# ---------------------------------------------------------------------------
# Воронка: асинхронный отчёт (create → poll → download CSV)
# ---------------------------------------------------------------------------
def _create_report(date_from: date, date_to: date) -> str:
    """POST создать отчёт воронки. download_id генерируем сами (uuid)."""
    download_id = str(uuid.uuid4())
    body = {
        "id": download_id,
        "reportType": "DETAIL_HISTORY_REPORT",
        "userReportName": "funnel_auto",
        "params": {
            "nmIDs": [],
            "subjectIds": [],
            "brandNames": [],
            "tagIds": [],
            "startDate": date_from.strftime("%Y-%m-%d"),
            "endDate": date_to.strftime("%Y-%m-%d"),
            "timezone": TIMEZONE,
            "aggregationLevel": AGG_LEVEL,
            "skipDeletedNm": False,
        },
    }
    resp = _request("POST", DOWNLOADS_URL, json_body=body)
    log.info("Отчёт воронки создан: id=%s, период %s..%s (ответ: %s)",
             download_id, date_from, date_to, resp.text[:80])
    return download_id


def _wait_report(download_id: str) -> None:
    """Опрос списка отчётов до status=SUCCESS для нашего id."""
    deadline = time.time() + POLL_TIMEOUT
    while time.time() < deadline:
        time.sleep(POLL_INTERVAL)
        data = _request("GET", DOWNLOADS_URL).json()
        items = data.get("data") or data
        if isinstance(items, dict):
            items = items.get("reports") or items.get("items") or []
        mine = [x for x in items if x.get("id") == download_id or x.get("downloadId") == download_id]
        if not mine:
            continue
        status = str(mine[0].get("status") or mine[0].get("state") or "").upper()
        if status in ("SUCCESS", "DONE", "OK", "READY"):
            return
        if status in ("FAILED", "ERROR"):
            raise RuntimeError(f"Отчёт воронки id={download_id} завершился ошибкой: {mine[0]}")
    raise TimeoutError(f"Отчёт воронки id={download_id} не готов за {POLL_TIMEOUT}s")


def _download_csv(download_id: str) -> list:
    """Скачивает ZIP отчёта, парсит CSV → list[dict] (ключи = заголовки CSV)."""
    resp = _request("GET", DOWNLOADS_FILE_URL.format(id=download_id), stream=True)
    content = resp.content
    texts = []
    if content[:2] == b"PK":
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for nm in zf.namelist():
                if nm.lower().endswith(".csv"):
                    texts.append(zf.read(nm).decode("utf-8-sig", errors="replace"))
    else:
        texts.append(content.decode("utf-8-sig", errors="replace"))

    records = []
    for txt in texts:
        # CSV воронки — разделитель запятая.
        reader = csv.DictReader(io.StringIO(txt), delimiter=",")
        if reader.fieldnames:
            log.info("Колонки CSV воронки: %s", reader.fieldnames)
        for row in reader:
            records.append(row)
    return records


def _num(v):
    if v in (None, ""):
        return ""
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return v


def fetch_funnel_day(target_date: date) -> dict:
    """Один день воронки → {nmID(str): запись CSV}. При нескольких строках берём первую."""
    download_id = _create_report(target_date, target_date)
    _wait_report(download_id)
    rows = _download_csv(download_id)
    by_nm = {}
    for r in rows:
        nm = str(r.get("nmID") or r.get("nmId") or "").strip()
        if nm and nm not in by_nm:
            by_nm[nm] = r
    log.info("Воронка за %s: строк=%d, уникальных nmID=%d", target_date, len(rows), len(by_nm))
    return by_nm


# ---------------------------------------------------------------------------
# Премиум-отчёт «поисковые запросы по товарам» — только рейтинги (Rating,
# FeedbackRating). Доступен при подписке Jam (у аккаунта проверено — есть).
# ---------------------------------------------------------------------------
def fetch_ratings_map(target_date: date) -> dict:
    """{nmID(str): {"rating":.., "feedbackRating":..}} из премиум-отчёта."""
    prev_day = target_date - timedelta(days=1)
    download_id = str(uuid.uuid4())
    body = {
        "id": download_id,
        "reportType": "SEARCH_QUERIES_PREMIUM_REPORT_PRODUCT",
        "userReportName": "ratings_auto",
        "params": {
            "currentPeriod": {"start": target_date.strftime("%Y-%m-%d"),
                              "end": target_date.strftime("%Y-%m-%d")},
            "pastPeriod": {"start": prev_day.strftime("%Y-%m-%d"),
                           "end": prev_day.strftime("%Y-%m-%d")},
            "nmIds": [],
            "orderBy": {"field": "openCard", "mode": "desc"},
            "positionCluster": "all",
            "includeSubstitutedSKUs": True,
            "includeSearchTexts": False,
        },
    }
    out = {}
    try:
        _request("POST", DOWNLOADS_URL, json_body=body)
        _wait_report(download_id)
        for r in _download_csv(download_id):
            nm = str(r.get("NmID") or r.get("nmID") or "").strip()
            if nm:
                out[nm] = {"rating": _num(r.get("Rating")),
                           "feedbackRating": _num(r.get("FeedbackRating"))}
        log.info("Премиум-отчёт: рейтинги для %d nmID", len(out))
    except Exception as e:
        log.warning("Премиум-отчёт (рейтинги) недоступен (%s) — рейтинги будут пустыми", e)
    return out


# ---------------------------------------------------------------------------
# Остатки на складах (асинхронный отчёт warehouse_remains)
# ---------------------------------------------------------------------------
def fetch_stocks_map() -> dict:
    """{nmID(str): {"wb":кол-во на складах WB, "mp":кол-во Маркетплейс(FBS)}}."""
    out = {}
    try:
        params = {"locale": "ru", "groupByNm": "true", "groupByBrand": "false",
                  "groupBySubject": "false", "groupBySa": "false",
                  "groupByBarcode": "false", "groupBySize": "false",
                  "filterPics": 0, "filterVolume": 0}
        # create
        headers = {"Authorization": WB_API_TOKEN}
        resp = requests.get(WAREHOUSE_REMAINS_URL, headers=headers, params=params, timeout=60)
        resp.raise_for_status()
        task_id = (resp.json().get("data") or {}).get("taskId")
        if not task_id:
            log.warning("warehouse_remains не вернул taskId — остатки будут пустыми")
            return out
        # poll
        deadline = time.time() + POLL_TIMEOUT
        while time.time() < deadline:
            time.sleep(POLL_INTERVAL)
            st = requests.get(WAREHOUSE_STATUS_URL.format(id=task_id), headers=headers, timeout=30)
            if str((st.json().get("data") or {}).get("status")).lower() == "done":
                break
        # download (большой JSON)
        dl = requests.get(WAREHOUSE_DOWNLOAD_URL.format(id=task_id), headers=headers, timeout=(30, 600))
        dl.raise_for_status()
        for row in dl.json():
            nm = str(row.get("nmId") or "").strip()
            if not nm:
                continue
            wb = mp = 0
            for w in row.get("warehouses") or []:
                if w.get("warehouseName") == WH_TOTAL_NAME:
                    wb = w.get("quantity") or 0
                elif w.get("warehouseName") == WH_MP_NAME:
                    mp = w.get("quantity") or 0
            out[nm] = {"wb": wb, "mp": mp}
        log.info("Остатки: nmID в отчёте: %d", len(out))
    except Exception as e:
        log.warning("Отчёт остатков недоступен (%s) — колонки остатков будут пустыми", e)
    return out


# ---------------------------------------------------------------------------
# Сборка строк листа "Товары"
# ---------------------------------------------------------------------------
def _ratio(a, b, digits=1):
    """a/b → округлённое число (доля/средняя цена). Пусто при делении на 0."""
    try:
        a = float(a or 0); b = float(b or 0)
        if b == 0:
            return ""
        v = a / b
        return round(v, digits) if digits else round(v)
    except (TypeError, ValueError):
        return ""


def build_row(nm: str, cur: dict, prev: dict, info: dict,
              rating: dict, stocks: dict, total_cur: float, total_prev: float) -> dict:
    """nmID + воронка (тек./пред.) + Content + рейтинги + остатки + вычисления → строка 'Товары'."""
    cur = cur or {}
    prev = prev or {}
    info = info or {}
    rating = rating or {}
    stocks = stocks or {}

    def c(key):   # значение текущего периода
        return _num(cur.get(key))

    def p(key):   # значение предыдущего периода
        return _num(prev.get(key))

    # Динамика суммы заказов = текущий - предыдущий
    dyn = ""
    try:
        dyn = float(cur.get("ordersSumRub") or 0) - float(prev.get("ordersSumRub") or 0)
    except (TypeError, ValueError):
        dyn = ""

    # Средняя цена = выручка / заказы (сошлось с браузерным файлом)
    avg_price_cur = _ratio(cur.get("ordersSumRub"), cur.get("ordersCount"), 0)
    avg_price_prev = _ratio(prev.get("ordersSumRub"), prev.get("ordersCount"), 0)
    # Доля карточки в выручке = выручка карточки / общая выручка * 100
    share_cur = _ratio((float(cur.get("ordersSumRub") or 0)) * 100, total_cur, 1)
    share_prev = _ratio((float(prev.get("ordersSumRub") or 0)) * 100, total_prev, 1)
    # Остатки и сумма остатков (кол-во * средняя цена)
    wb_qty = stocks.get("wb", "")
    mp_qty = stocks.get("mp", "")
    stock_sum = ""
    if isinstance(avg_price_cur, (int, float)):
        total_qty = (wb_qty or 0) + (mp_qty or 0)
        stock_sum = round(total_qty * avg_price_cur) if total_qty else ""

    return {
        "Артикул продавца": info.get("vendorCode", ""),
        "Артикул WB": _num(nm),
        "Название": info.get("title", ""),
        "Предмет": info.get("subjectName", ""),
        "Бренд": info.get("brand", ""),
        "Удаленный товар": "Нет" if info else "",
        "Рейтинг карточки": rating.get("rating", ""),
        "Рейтинг по отзывам": rating.get("feedbackRating", ""),
        "Показы": "",                          # нет в API (рекламный/поисковый источник)
        "Показы (предыдущий период)": "",
        "Доля карточки в выручке": share_cur,
        "Доля карточки в выручке (предыдущий период)": share_prev,
        "Переходы в карточку": c("openCardCount"),
        "Переходы в карточку (предыдущий период)": p("openCardCount"),
        "Положили в корзину": c("addToCartCount"),
        "Положили в корзину (предыдущий период)": p("addToCartCount"),
        "Заказали товаров, шт": c("ordersCount"),
        "Заказали товаров, шт (предыдущий период)": p("ordersCount"),
        "Выкупили, шт": c("buyoutsCount"),
        "Выкупы, шт (предыдущий период)": p("buyoutsCount"),
        "Отменили, шт": c("cancelCount"),
        "Отменили, шт (предыдущий период)": p("cancelCount"),
        "Конверсия в корзину, %": c("addToCartConversion"),
        "Конверсия в корзину, % (предыдущий период)": p("addToCartConversion"),
        "Конверсия в заказ, %": c("cartToOrderConversion"),
        "Конверсия в заказ, % (предыдущий период)": p("cartToOrderConversion"),
        "Процент выкупа": c("buyoutPercent"),
        "Процент выкупа (предыдущий период)": p("buyoutPercent"),
        "Заказали на сумму, ₽": c("ordersSumRub"),
        "Заказали на сумму, ₽ (предыдущий период)": p("ordersSumRub"),
        "Динамика суммы заказов, ₽": dyn,
        "Выкупили на сумму, ₽": c("buyoutsSumRub"),
        "Выкупили на сумму, ₽ (предыдущий период)": p("buyoutsSumRub"),
        "Отменили на сумму, ₽": c("cancelSumRub"),
        "Отменили на сумму, ₽ (предыдущий период)": p("cancelSumRub"),
        "Средняя цена, ₽": avg_price_cur,
        "Средняя цена, ₽ (предыдущий период)": avg_price_prev,
        # Период = 1 день, поэтому среднее заказов/день = число заказов.
        "Среднее количество заказов в день, шт": c("ordersCount"),
        "Среднее количество заказов в день, шт (предыдущий период)": p("ordersCount"),
        "Остатки «Склад WB», шт": wb_qty,
        "Остатки МП, шт": mp_qty,
        "Сумма остатков на складах, ₽": stock_sum,
        "Среднее время доставки": "",          # нет в API воронки
        "Среднее время доставки (предыдущий период)": "",
        "Локальные заказы, %": "",             # нет в API воронки
        "Локальные заказы, % (предыдущий период)": "",
    }


def write_workbook(rows: list, dest: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.cell(row=1, column=1, value=SHEET_TITLE)
    for j, col in enumerate(TOVARY_COLUMNS, start=1):
        ws.cell(row=2, column=j, value=col)
    for i, rec in enumerate(rows, start=3):
        for j, col in enumerate(TOVARY_COLUMNS, start=1):
            val = rec.get(col, "")
            ws.cell(row=i, column=j, value=(None if val == "" else val))
    wb.save(dest)


def _wb_filename(target_date: date) -> str:
    """Имя как у браузерной выгрузки: 'DD.MM с D-M-YYYY по D-M-YYYY.xlsx'."""
    prefix = target_date.strftime("%d.%m")
    rng = f"{target_date.day}-{target_date.month}-{target_date.year}"
    return f"{prefix} с {rng} по {rng}.xlsx"


def download_wb_funnel_report(target_date: date = None, output_dir: Path = None) -> bool:
    log.info("Starting WB Funnel API Downloader...")

    if not WB_API_TOKEN:
        msg = "WB_API_TOKEN не задан — нужен токен WB (категории «Аналитика» и «Контент»). Добавьте в .env."
        log.error(msg)
        print(f"[ERROR] {msg}")
        return False

    if target_date is None:
        target_date = date.today()

    DOWNLOADS_DIR = Path(output_dir) if output_dir else DEFAULT_DOWNLOADS_DIR
    if output_dir:
        DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    else:
        try:
            DEFAULT_DOWNLOADS_DIR.stat()
        except Exception as e:
            print(f"[ERROR] Cannot access downloads folder: {downloads_folder}. {e}")
            return False

    log.info("Target date (период): %s", target_date)

    try:
        # Справочник карточек (артикул/название/предмет/бренд)
        content = fetch_content_map()

        # Воронка: текущий и предыдущий день
        cur = fetch_funnel_day(target_date)
        prev = fetch_funnel_day(target_date - timedelta(days=1))

        if not cur:
            log.error("Воронка вернула 0 строк за %s", target_date)
            print(f"[ERROR] Нет данных воронки WB за {target_date}")
            return False

        # Обогащение: рейтинги (премиум-отчёт) и остатки (warehouse_remains)
        ratings = fetch_ratings_map(target_date)
        stocks = fetch_stocks_map()

        # Итоги выручки за оба периода — для «Доли карточки в выручке»
        def _sum_rev(d):
            s = 0.0
            for r in d.values():
                try:
                    s += float(r.get("ordersSumRub") or 0)
                except (TypeError, ValueError):
                    pass
            return s
        total_cur = _sum_rev(cur)
        total_prev = _sum_rev(prev)

        rows = [build_row(nm, cur.get(nm), prev.get(nm), content.get(nm),
                          ratings.get(nm), stocks.get(nm), total_cur, total_prev)
                for nm in cur]
        dest = DOWNLOADS_DIR / _wb_filename(target_date)
        write_workbook(rows, dest)
        log.info("SUCCESS: сохранено %d строк (лист 'Товары') → %s", len(rows), dest)
        print(f"SUCCESS: saved {len(rows)} rows -> {dest}")
        return True

    except Exception as e:
        log.exception("Ошибка выгрузки воронки WB: %s", e)
        print(f"[ERROR] {e}")
        return False


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Download Wildberries sales-funnel report (sheet 'Товары') via API")
    parser.add_argument("--date", type=str, default=None,
                        help="Дата данных YYYY-MM-DD (по умолчанию сегодня)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Куда сохранить файл (по умолчанию сетевая шара Taldykin)")
    args = parser.parse_args()

    target = None
    if args.date:
        target = datetime.strptime(args.date, "%Y-%m-%d").date()

    success = download_wb_funnel_report(target_date=target, output_dir=args.output_dir)
    sys.stdout.flush(); sys.stderr.flush()  # os._exit не сбрасывает буферы print()
    os._exit(0 if success else 1)
