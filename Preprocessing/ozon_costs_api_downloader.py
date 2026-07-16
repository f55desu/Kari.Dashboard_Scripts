"""
ozon_costs_api_downloader.py — Выгружает затраты на рекламу Ozon через
Performance API и формирует xlsx с листами Statistics и Union (как файл в
   ...\Затраты\Озон. Затраты из Аналитики New Format\Аналитика продвижения_DD.MM.YYYY.xlsx)

⚠️  РЕЖИМ: CAMPAIGN-LEVEL (по кампаниям), не per-SKU.
    Причина: per-SKU статистику Ozon отдаёт только через async-отчёт
    /api/client/statistics, а он ограничен «максимум 10 кампаний на отчёт» и
    «1 отчёт одновременно» + дневная квота. На аккаунте с тысячами активных
    кампаний это упирается в 403 и практически недостижимо. Поэтому данные
    берём из СИНХРОННОГО эндпоинта (без квоты и без пагинационного капа):
        GET /api/client/statistics/daily/json?dateFrom=&dateTo=
    Он возвращает JSON по КАЖДОЙ кампании с активностью за период:
        {"rows":[{"id","title","date","views","clicks","moneySpent",
                  "orders","ordersMoney"}]}
    (Синхронный /campaign/product НЕ используется: он жёстко отдаёт только
    первые 1000 кампаний по возрастанию ID — активные туда не попадают.)
    Метрик Показы/Клики/Расход/Продано/Продажи достаточно; CTR, Ср. стоимость
    клика, ДРР, Затраты на заказ ВЫЧИСЛЯЮТСЯ. «Добавления в корзину» в daily/json
    нет — колонка остаётся пустой.

Лист Statistics (16 колонок, период в строке 1, заголовки во 2-й):
    SKU | Название товара | Инструмент | Место размещения | ID кампании |
    Расход, ₽ | ДРР в продвижении, % | Продажи в продвижении, ₽ |
    Продано товаров, шт | CTR, % | Показы | Клики | Добавления в корзину, шт |
    Конверсия в корзину, % | Затраты на заказ, ₽ | Стоимость клика, ₽
  В campaign-level режиме:
    • «Название товара» = название кампании (нет разбивки по товарам);
    • SKU / «Конверсия в корзину, %» — пустые (нет в источнике);
    • «Затраты на заказ, ₽» = Расход / Продано товаров (вычисляется).

Лист Union (6 колонок) — per-SKU атрибуция; через синхронный API недоступна,
поэтому лист создаётся со структурой (период + заголовки), без строк данных.

Креды Performance API (OAuth client_credentials) из .env:
    OZON_PERF_CLIENT_ID / OZON_PERF_CLIENT_SECRET

Запуск:
    python ozon_costs_api_downloader.py [--output-dir PATH] [--date YYYY-MM-DD]
  --date        дата данных (по умолчанию вчера). Период = [date, date].
  --output-dir  куда сохранить (по умолчанию сетевая шара Taldykin).
"""
import sys
import os
import io
import csv
import json
import time
import logging
from datetime import date, timedelta, datetime
from pathlib import Path

import requests
from openpyxl import Workbook

if sys.stdout is not None:
    sys.stdout.reconfigure(encoding="utf-8")
if sys.stderr is not None:
    sys.stderr.reconfigure(encoding="utf-8")

PROJECT_DIR = Path(__file__).parent
try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_DIR / ".env")
except Exception:
    pass

downloads_folder = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Taldykin"
DEFAULT_DOWNLOADS_DIR = Path(downloads_folder)

LOG_FILE = os.path.join(PROJECT_DIR, "ozon_costs_api.log")
logging.basicConfig(
    filename=LOG_FILE, level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8',
)
log = logging.getLogger(__name__)

# --- Ozon Performance API ---
PERF_BASE = "https://api-performance.ozon.ru"
TOKEN_URL = f"{PERF_BASE}/api/client/token"
CAMPAIGN_URL = f"{PERF_BASE}/api/client/campaign"
DAILY_JSON_URL = f"{PERF_BASE}/api/client/statistics/daily/json"

OZON_PERF_CLIENT_ID = os.environ.get("OZON_PERF_CLIENT_ID", "").strip()
OZON_PERF_CLIENT_SECRET = os.environ.get("OZON_PERF_CLIENT_SECRET", "").strip()

REQUEST_PAUSE = 0.3
MAX_RETRIES = 6
RETRY_STATUS = {429, 500, 502, 503, 504}

STAT_COLUMNS = [
    "SKU", "Название товара", "Инструмент", "Место размещения", "ID кампании",
    "Расход, ₽", "ДРР в продвижении, %", "Продажи в продвижении, ₽",
    "Продано товаров, шт", "CTR, %", "Показы", "Клики", "Добавления в корзину, шт",
    "Конверсия в корзину, %", "Затраты на заказ, ₽", "Стоимость клика, ₽",
]
UNION_COLUMNS = [
    "SKU в продвижении", "Название товара в продвижении",
    "SKU из объединенной карточки", "Название товара из объединенной карточки",
    "Продажи, ₽", "Заказы, шт",
]

# PaymentType кампании → «Инструмент».
PAYMENT_TYPE_LABELS = {
    "CPC": "Оплата за клик",
    "CPO": "Оплата за заказ",
    "CPM": "Оплата за показы",
}
# placement кампании (значения массива placement из /api/client/campaign) →
# «Место размещения» как в оригинале.
PLACEMENT_LABELS = {
    "PLACEMENT_SEARCH_AND_CATEGORY": "Поиск и рекомендации",
    "PLACEMENT_PDP": "Карточка товара",
    "PLACEMENT_TOP": "Вывод в топ",
    "PLACEMENT_TOP_PROMOTION": "Вывод в топ",
    "PLACEMENT_OVERTOP": "Поиск и рекомендации",
}
# /api/client/campaign отдаёт не все кампании (старые активные могут
# отсутствовать). Когда кампании нет в справочнике — берём доминирующие на
# аккаунте значения (94% кампаний: CPC + поиск/категории), как в браузерном файле.
DEFAULT_PT = "CPC"
DEFAULT_PLACEMENT = "PLACEMENT_SEARCH_AND_CATEGORY"

_token_cache = {"access_token": None, "expires_at": 0.0}


def _request(method: str, url: str, *, token: str = None, json_body: dict = None,
             params: dict = None) -> requests.Response:
    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    attempt = 0
    while True:
        attempt += 1
        try:
            resp = requests.request(method, url, headers=headers,
                                    data=json.dumps(json_body) if json_body is not None else None,
                                    params=params, timeout=120)
        except requests.RequestException as e:
            if attempt >= MAX_RETRIES:
                raise
            wait = min(60, 5 * attempt)
            log.warning("Сетевая ошибка (%s); попытка %d/%d, пауза %ds", e, attempt, MAX_RETRIES, wait)
            time.sleep(wait)
            continue
        if resp.status_code in RETRY_STATUS:
            if attempt >= MAX_RETRIES:
                resp.raise_for_status()
            wait = min(30, 3 * attempt)
            log.warning("HTTP %d (%s); попытка %d/%d, пауза %ds",
                        resp.status_code, resp.text[:120], attempt, MAX_RETRIES, wait)
            time.sleep(wait)
            continue
        resp.raise_for_status()
        time.sleep(REQUEST_PAUSE)
        return resp


def get_token() -> str:
    now = time.time()
    if _token_cache["access_token"] and now < _token_cache["expires_at"] - 60:
        return _token_cache["access_token"]
    data = _request("POST", TOKEN_URL, json_body={
        "client_id": OZON_PERF_CLIENT_ID,
        "client_secret": OZON_PERF_CLIENT_SECRET,
        "grant_type": "client_credentials",
    }).json()
    token = data.get("access_token")
    if not token:
        raise RuntimeError(f"Не получен access_token Performance API: {data}")
    _token_cache["access_token"] = token
    _token_cache["expires_at"] = now + float(data.get("expires_in", 1800))
    log.info("Получен токен Performance API")
    return token


CAMPAIGN_IDS_BATCH = 500   # /api/client/campaign?campaignIds=... принимает ≥500 id


def fetch_campaign_meta(token: str, ids) -> dict:
    """{campaignId(str): {"pt": PaymentType, "placement": первый placement}}.

    Тянем метаданные ТОЛЬКО по нужным кампаниям (id из daily/json) через фильтр
    campaignIds батчами. Так достаются и старые активные кампании, которых нет
    в общем листинге /api/client/campaign без фильтра.
    """
    out = {}
    ids = [str(i) for i in ids]
    for i in range(0, len(ids), CAMPAIGN_IDS_BATCH):
        batch = ids[i:i + CAMPAIGN_IDS_BATCH]
        try:
            camps = _request("GET", CAMPAIGN_URL, token=token,
                             params=[("campaignIds", c) for c in batch]).json().get("list") or []
        except Exception as e:
            log.warning("Метаданные кампаний, батч %d: %s (пропускаем)", i, e)
            continue
        for c in camps:
            placements = c.get("placement") or []
            out[str(c.get("id"))] = {
                "pt": c.get("PaymentType") or c.get("paymentType") or "",
                "placement": placements[0] if placements else "",
            }
    log.info("Метаданных кампаний получено: %d из %d запрошенных", len(out), len(ids))
    return out


def fetch_daily(token: str, target_date: date) -> dict:
    """Синхронный /daily/json за [date, date] → {campaignId: агрегат метрик}.

    Возвращает все кампании с активностью (без пагинационного капа). При периоде
    в один день на кампанию одна строка; на всякий случай суммируем по id.
    """
    day = target_date.strftime("%Y-%m-%d")
    data = _request("GET", DAILY_JSON_URL, token=token,
                    params={"dateFrom": day, "dateTo": day}).json()
    rows = data.get("rows") or []
    agg = {}
    for r in rows:
        cid = str(r.get("id"))
        a = agg.setdefault(cid, {"title": r.get("title", ""), "views": 0.0, "clicks": 0.0,
                                 "moneySpent": 0.0, "orders": 0.0, "ordersMoney": 0.0})
        for src, key in (("views", "views"), ("clicks", "clicks"), ("moneySpent", "moneySpent"),
                         ("orders", "orders"), ("ordersMoney", "ordersMoney")):
            v = _num(r.get(src))
            if isinstance(v, (int, float)):
                a[key] += v
    log.info("daily/json: строк=%d, уникальных кампаний=%d", len(rows), len(agg))
    return agg


def _num(value):
    """'1 029,83' / '0' / '' → float|int|''. Русская запятая → точка."""
    if value in (None, ""):
        return ""
    s = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return value


def _n(x):
    """float|int|'' → нормализованное число для записи (int если целое)."""
    try:
        f = float(x)
        return int(f) if f.is_integer() else round(f, 6)
    except (TypeError, ValueError):
        return ""


def build_stat_rows(daily: dict, meta: dict) -> list:
    """daily/json (агрегат по кампаниям) + метаданные → строки листа Statistics."""
    rows = []
    for cid, a in daily.items():
        views, clicks = a["views"], a["clicks"]
        spent, orders, rev = a["moneySpent"], a["orders"], a["ordersMoney"]
        if not (spent or views or clicks):
            continue  # пропускаем нулевые
        m = meta.get(cid, {})
        ctr = (clicks / views * 100) if views else ""
        cpc = (spent / clicks) if clicks else ""
        drr = (spent / rev * 100) if rev else ""
        cpo = (spent / orders) if orders else ""
        rows.append({
            "SKU": "",  # campaign-level: разбивки по SKU нет
            "Название товара": a.get("title", ""),   # здесь — название кампании
            "Инструмент": PAYMENT_TYPE_LABELS.get(m.get("pt"), m.get("pt", "")),
            "Место размещения": PLACEMENT_LABELS.get(m.get("placement"), m.get("placement", "")),
            "ID кампании": _n(cid),
            "Расход, ₽": _n(spent),
            "ДРР в продвижении, %": _n(drr),
            "Продажи в продвижении, ₽": _n(rev),
            "Продано товаров, шт": _n(orders),
            "CTR, %": _n(ctr),
            "Показы": _n(views),
            "Клики": _n(clicks),
            "Добавления в корзину, шт": "",   # нет в daily/json
            "Конверсия в корзину, %": "",     # нет в daily/json
            "Затраты на заказ, ₽": _n(cpo),
            "Стоимость клика, ₽": _n(cpc),
        })
    return rows


def _write_sheet(ws, columns, rows, period_title):
    ws.cell(row=1, column=1, value=period_title)
    for j, col in enumerate(columns, start=1):
        ws.cell(row=2, column=j, value=col)
    for i, rec in enumerate(rows, start=3):
        for j, col in enumerate(columns, start=1):
            val = rec.get(col, "")
            ws.cell(row=i, column=j, value=(None if val == "" else val))


def write_workbook(stat_rows, target_date, dest):
    period = f"Период: {target_date.strftime('%d.%m.%Y')} - {target_date.strftime('%d.%m.%Y')}"
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistics"
    _write_sheet(ws, STAT_COLUMNS, stat_rows, period)
    ws2 = wb.create_sheet("Union")
    _write_sheet(ws2, UNION_COLUMNS, [], period)  # per-SKU атрибуция недоступна через sync API
    wb.save(dest)


def download_ozon_costs_report(target_date: date = None, output_dir: Path = None) -> bool:
    log.info("Starting Ozon Costs (Performance API, campaign-level) Downloader...")

    if not OZON_PERF_CLIENT_ID or not OZON_PERF_CLIENT_SECRET:
        msg = "OZON_PERF_CLIENT_ID / OZON_PERF_CLIENT_SECRET не заданы. Добавьте в .env."
        log.error(msg)
        print(f"[ERROR] {msg}")
        return False

    if target_date is None:
        target_date = date.today() - timedelta(days=1)

    DOWNLOADS_DIR = Path(output_dir) if output_dir else DEFAULT_DOWNLOADS_DIR
    if output_dir:
        DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    else:
        try:
            DEFAULT_DOWNLOADS_DIR.stat()
        except Exception as e:
            print(f"[ERROR] Cannot access downloads folder: {downloads_folder}. {e}")
            return False

    log.info("Target date: %s", target_date)
    try:
        token = get_token()
        daily = fetch_daily(token, target_date)
        meta = fetch_campaign_meta(token, daily.keys())
        stat_rows = build_stat_rows(daily, meta)
        log.info("Statistics строк (кампаний с активностью): %d", len(stat_rows))

        file_date = target_date + timedelta(days=1)
        filename = f"Аналитика продвижения_{file_date.strftime('%d.%m.%Y')}.xlsx"
        dest = DOWNLOADS_DIR / filename
        write_workbook(stat_rows, target_date, dest)
        log.info("SUCCESS: Statistics=%d (Union пуст — per-SKU недоступен) → %s", len(stat_rows), dest)
        print(f"SUCCESS: saved Statistics={len(stat_rows)} rows (campaign-level), Union=0 -> {dest}")
        return True
    except Exception as e:
        log.exception("Ошибка выгрузки затрат Ozon: %s", e)
        print(f"[ERROR] {e}")
        return False


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Download Ozon ad-costs (campaign-level) via Performance API")
    parser.add_argument("--date", type=str, default=None, help="Дата данных YYYY-MM-DD (по умолчанию вчера)")
    parser.add_argument("--output-dir", type=str, default=None, help="Куда сохранить файл")
    args = parser.parse_args()

    target = None
    if args.date:
        target = datetime.strptime(args.date, "%Y-%m-%d").date()

    success = download_ozon_costs_report(target_date=target, output_dir=args.output_dir)
    sys.stdout.flush(); sys.stderr.flush()
    os._exit(0 if success else 1)
