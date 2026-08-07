import os
import re
import shutil
import zipfile
from datetime import datetime, timedelta
import xlwings as xw
import time

from openpyxl import load_workbook
import pandas as pd

import win32com.client as win32

# 🔧 Пути
downloads_folder = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Taldykin"
current_folder = os.getcwd()

folder_wb = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\ВЫГРУЗКА воронка ВБ"
folder_zatraty_wb = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\Затраты\Затраты ВБ"
folder_week = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!\Показатели по неделям ВБ"
folder_campaign_info = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Федоров\Дашбоард по рекламным кампаниям\!!!_ИСХОДНИКИ ДЛЯ ДАШБОРДА_НЕ УДАЛЯТЬ_!!!"

# folder_wb = "ВЫГРУЗКА воронка ВБ"
# folder_zatraty_wb = "Затраты\\Затраты ВБ"
# folder_week = "Показатели по неделям ВБ"
# folder_campaign_info = current_folder

file_voronka = ''
query_name = "ВЫГРУЗКА воронка ВБ"

def extract_dates_from_query(wb_path, query_name):
    app = xw.App(visible=False)
    try:
        wb = app.books.open(wb_path) 
        xl = wb.app.api
        xl_wb = xl.Workbooks(os.path.basename(wb_path))

        # Найти запрос
        for q in xl_wb.Queries:
            if q.Name == query_name:
                m_code = q.Formula
                break
        else:
            raise ValueError(f"Запрос '{query_name}' не найден")

        # Найти все вхождения Text.Contains([Name], "дата")
        date_matches = re.findall(r'Text\.Contains\(\[Name\],\s*"(\d{1,2}-\d{1,2}-\d{4})"\)', m_code)

        # Удалить дубликаты и отсортировать
        unique_sorted_dates = sorted(set(date_matches), key=lambda d: datetime.strptime(d, "%d-%m-%Y"))
        print(f"📆 Найдены даты в фильтре: {unique_sorted_dates}")
        return unique_sorted_dates
    finally:
        app.quit()


# 📦 Этап 1: Распаковать архив и переименовать файл
def extract_latest_vb_file():
    zips = [f for f in os.listdir(downloads_folder) if f.endswith('.zip') and re.match(r"\d{2}\.\d{2} с \d{2}\.\d{2}\.\d{4} по \d{2}\.\d{2}\.\d{4}\.zip", f)]
    if not zips:
        raise FileNotFoundError("Нет zip-файлов нужного формата в Загрузках")
    latest_zip = max(zips, key=lambda x: os.path.getmtime(os.path.join(downloads_folder, x)))
    zip_path = os.path.join(downloads_folder, latest_zip)
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        extracted_files = zip_ref.namelist()
        xlsx_files = [f for f in extracted_files if f.endswith('.xlsx')]
        if not xlsx_files:
            raise FileNotFoundError("Нет .xlsx файлов в архиве")
        latest_xlsx = max(xlsx_files, key=lambda f: os.path.getmtime(zip_ref.extract(f, downloads_folder)))

    # Переименование и перемещение
    src = os.path.join(downloads_folder, latest_xlsx)
    
    # Преобразование имени: 01.02 с 01.02.2026 по 01.02.2026.zip -> 01.02 с 1-2-2026 по 1-2-2026.xlsx
    # Извлекаем части из имени zip-файла
    match = re.search(r"(\d{2}\.\d{2}) с (\d{2})\.(\d{2})\.(\d{4}) по (\d{2})\.(\d{2})\.(\d{4})\.zip", latest_zip)
    if not match:
        raise ValueError("Невозможно извлечь даты из имени zip-файла")
    
    prefix = match.group(1)  # 01.02
    day1, month1, year1 = int(match.group(2)), int(match.group(3)), int(match.group(4))  # убираем ведущие нули через int()
    day2, month2, year2 = int(match.group(5)), int(match.group(6)), int(match.group(7))
    
    # Формируем новое имя файла
    new_name = f"{prefix} с {day1}-{month1}-{year1} по {day2}-{month2}-{year2}.xlsx"
    
    dst = os.path.join(folder_wb, new_name)
    shutil.move(src, dst)
    file_voronka = new_name
    print(f"✅ Этап 1: файл {new_name} перемещён в ВЫГРУЗКА воронка ВБ")
    return os.path.join(folder_wb, new_name)

# 💰 Этап 2: Копировать файл История-затрат
def copy_latest_cost_history():
    # Новый паттерн для поиска файла "Не определено"
    pattern = r"История-затрат-Не определено-.*\.xlsx"
    
    # Ищем файлы, соответствующие паттерну
    files = [f for f in os.listdir(downloads_folder) if re.match(pattern, f)]
    
    # Если файлы не найдены, выбрасываем ошибку
    if not files:
        raise FileNotFoundError("Файл История-затрат не найден")
    
    # Находим последний файл по времени изменения
    latest = max(files, key=lambda f: os.path.getmtime(os.path.join(downloads_folder, f)))

    # Путь к последнему найденному файлу
    latest_file_path = os.path.join(downloads_folder, latest)

    # Переименование файла в новый паттерн "Все"
    renamed_file = re.sub(r"История-затрат-Не определено", "История-затрат-Все", latest)
    renamed_file_path = os.path.join(downloads_folder, renamed_file)

    # Переименовываем файл
    os.rename(latest_file_path, renamed_file_path)
    print(f"✅ Файл переименован в: {renamed_file}")

    # Очистка целевой папки
    for f in os.listdir(folder_zatraty_wb):
        os.remove(os.path.join(folder_zatraty_wb, f))

    # Копирование переименованного файла в целевую папку
    shutil.copy2(renamed_file_path, folder_zatraty_wb)
    
    print(f"✅ Этап 2: файл {renamed_file} скопирован в Затраты ВБ и папка очищена")
    
def unify_columns(filepath, ethalon=os.path.join(downloads_folder, 'Воронка ВБ Эталон.xlsx'), sheet_name='Товары', header_row=2):
    """Приводит файл к структуре эталона: удаляет лишние столбцы И добавляет
       отсутствующие из эталона как пустые (см. 07.05.2026 — ВБ убрал «Остатки склад ВБ, шт»
       и «Остатки МП, шт», из-за чего Power Query терял данные за свежие дни)."""
    ethalon_df = pd.read_excel(ethalon, sheet_name=sheet_name, header=1)
    ethalon_headers = ethalon_df.columns.tolist()

    wb = load_workbook(filename=filepath)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[header_row]]

    # 1) Удаляем столбцы, которых нет в эталоне
    columns_to_delete = [h for h in headers if h not in ethalon_headers]
    for col in reversed(columns_to_delete):
        col_index = headers.index(col) + 1
        ws.delete_cols(col_index)
    headers = [cell.value for cell in ws[header_row]]

    # 2) Добавляем отсутствующие столбцы из эталона (на нужные позиции)
    for target_idx, eth_col in enumerate(ethalon_headers, start=1):
        if eth_col not in headers:
            ws.insert_cols(target_idx)
            ws.cell(row=header_row, column=target_idx, value=eth_col)
            headers.insert(target_idx - 1, eth_col)
            print(f"  ➕ Добавлен отсутствующий столбец '{eth_col}' (позиция {target_idx})")

    # 3) "." → "," в числовых строках
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            val = cell.value
            if isinstance(val, str):
                try:
                    float(val.replace(',', '.'))
                    if '.' in val:
                        cell.value = val.replace('.', ',')
                except ValueError:
                    pass

    wb.save(filepath)
    print(f"✅ Столбцы унифицированы (удалено лишних: {len(columns_to_delete)}): {filepath}")

# Этап 3: Копировать файл История-затрат
def duplicate_latest_week_file():
    pattern = r"Показы и затраты_(\d+)\sнеделя\.xlsx"
    files = [f for f in os.listdir(folder_week) if re.match(pattern, f)]
    if not files:
        raise FileNotFoundError("Файлы 'Показы и затраты' не найдены")

    def extract_week(f): return int(re.search(pattern, f).group(1))
    latest_file = max(files, key=extract_week)
    latest_week = extract_week(latest_file)
    new_week = latest_week + 1

    new_filename = f"Показы и затраты_{new_week} неделя.xlsx"
    src_path = os.path.join(folder_week, latest_file)
    dst_path = os.path.join(folder_week, new_filename)

    shutil.copy2(src_path, dst_path)
    print(f"✅ Этап 3 Скопирован файл: {latest_file} → {new_filename}")
    return dst_path

# 🧠 Шаг 3.1–3.3: Обновить шаги в Power Query
def update_power_query_steps(wb_path, query_name, old_dates):
    # ➕ Вычислить новые даты (+7 дней)
    start_date = datetime.now() - timedelta(days=1)
    new_dates = [
        # (datetime.strptime(d, "%d-%m-%Y") + timedelta(days=7)).strftime("%d-%m-%Y")
        # for d in old_dates
        f"{(start_date + timedelta(days=i)).day}-{(start_date + timedelta(days=i)).month}-{(start_date + timedelta(days=i)).year}"
        for i in range(7)
    ]

    app = xw.App(visible=False)
    try:
        wb = xw.Book(wb_path)
        xl = wb.app.api
        xl_wb = xl.Workbooks(os.path.basename(wb_path))

        # Найти запрос
        for q in xl_wb.Queries:
            if q.Name == query_name:
                m_code = q.Formula
                break
        else:
            raise ValueError(f"Запрос '{query_name}' не найден")

        # Шаг 3.1: Заменить все Text.Contains по старым датам на новые
        for old, new in zip(old_dates, new_dates):
            m_code = m_code.replace(f'Text.Contains([Name], "{old}")', f'Text.Contains([Name], "{new}")')

        # Шаги 3.2-3.3: Заменить имя в шаге "Переименованные столбцы"
        # 🔽 Очистка заголовка из A1
        # yesterday = datetime.now() + timedelta(weeks=1) - timedelta(days=1) # TEST
        yesterday = datetime.now() - timedelta(days=1) # RIGHT VERSION
        day = yesterday.day
        month = yesterday.month
        year = yesterday.year
        new_filename = f'{yesterday.strftime("%d.%m")} с {day}-{month}-{year} по {day}-{month}-{year}.xlsx'

        # 🧠 Найти текущее имя файла в M-коде
        old_filename_match = re.search(
            r'"(\d{2}\.\d{2} с \d{1,2}-\d{1,2}-\d{4} по \d{1,2}-\d{1,2}-\d{4}\.xlsx)"',
            m_code
        )

        if not old_filename_match:
            raise ValueError("Не найдено старое имя файла в шаге 'Переименованные столбцы'")

        old_filename = old_filename_match.group(1)

        # 🔁 Безопасная замена
        m_code = m_code.replace(f'"{old_filename}"', f'"{new_filename}"')
        print(f"🔁 Заменено имя файла: {old_filename} → {new_filename}")

        # old_filename_match = re.search(r'"(\d{2}\.\d{2} с \d{1,2}-\d{1,2}-\d{4} по \d{1,2}-\d{1,2}-\d{4}\.xlsx)"', m_code)
        # if not old_filename_match:
        #     raise ValueError("Не найдено старое имя файла в M-коде")
        # old_filename = old_filename_match.group(1)
        # m_code = m_code.replace(old_filename, f'"{new_filename_header}"')

        # Применить изменения
        q.Formula = m_code
        print("🔁 Этапы 3.1-3.3 Шаги запроса обновлены")

        # Шаг 4: Обновить и сохранить
        for conn in xl_wb.Connections:
            if conn.Name == query_name:
                conn.Refresh()
                break

        wb.save()
        wb.close()
    finally:
        app.quit()

# 📈 Этап 4: Найти файл с максимальной неделей и обновить
def update_latest_week_file():
    pattern = r"Показы и затраты_(\d+)\sнеделя\.xlsx"
    week_files = [f for f in os.listdir(folder_week) if re.match(pattern, f)]
    if not week_files:
        raise FileNotFoundError("Файлы 'Показы и затраты' не найдены")

    def extract_week(f): return int(re.search(pattern, f).group(1))
    latest_week_file = max(week_files, key=extract_week)

    full_path = os.path.abspath(os.path.join(folder_week, latest_week_file))
    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Файл не найден: {full_path}")

    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False  # Отключает предупреждения Excel
    try:
        wb = excel.Workbooks.Open(full_path)
        wb.RefreshAll()
        wb.Save()
        wb.Close()
        print(f"✅ Этап 4: файл '{latest_week_file}' обновлён")
    finally:
        excel.Quit()

# 📤 Этап 5: Переместить статусы_отчет.xlsx → Campaign_Info.xlsx
def move_campaign_info():
    src_file = os.path.join(downloads_folder, "статусы_отчет.xlsx")
    dst_file = os.path.join(folder_campaign_info, "Campaign_Info.xlsx")
    if not os.path.exists(src_file):
        raise FileNotFoundError("Файл статусы_отчет.xlsx не найден")
    shutil.move(src_file, dst_file)
    print("✅ Этап 5: Campaign_Info.xlsx перемещён в рабочую папку")

# 📊 Этап 6: Обновить Затраты ВБ_2.xlsx
def update_zatraty_wb2():
    file = os.path.join(folder_campaign_info, "Затраты ВБ_2.xlsx")
    if not os.path.exists(file):
        raise FileNotFoundError("Файл Затраты ВБ_2.xlsx не найден в рабочей папке")

    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False  # Отключает предупреждения Excel
    try:
        wb = excel.Workbooks.Open(file)
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        time.sleep(5)

        sheet = wb.Sheets["Лист1"]
        for pivot in sheet.PivotTables():
            pivot.RefreshTable()
        
        time.sleep(5)  # ⏳ Подождать 5 секунд
        # excel.CalculateUntilAsyncQueriesDone()  # Дожидаемся завершения обновления
        wb.Save()
        print("✅ Этап 6: Затраты ВБ_2.xlsx обновлён")
    finally:
        wb.Close()
        excel.Quit()


def main():
    """Main execution function for Dashboard WB Wrapper (Tuesday version)"""
    logs = open('logs.log', 'a')
    logs.write(f'{datetime.now()} - WB Dashboard Wrapper ran\n')

    voronka_wb = extract_latest_vb_file() # 1
    unify_columns(voronka_wb)
    copy_latest_cost_history() # 2
    # # 3
    # new_file_path = duplicate_latest_week_file()
    # old_dates = extract_dates_from_query(new_file_path, query_name)
    # update_power_query_steps(new_file_path, query_name, old_dates) # 3.1-3.3

    # update_latest_week_file() # 4
    # update_zatraty_wb2() # 6
    
    # Запускаем скрипт
    # subprocess.run(['DB_WB_1.1.exe'], check=True)

    logs = open('logs.log', 'a')
    logs.write(f'{datetime.now()} - WB Dashboard Wrapper completed\n')
    logs.close()

# ▶️ Запуск всех этапов
if __name__ == "__main__":
    main()
