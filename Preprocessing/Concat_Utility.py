import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# 🔧 Пути
downloads_folder = r"\\kari.local\public\all\Analytics\Marketplaceanalytics\Taldykin"
current_folder = os.getcwd()

# ✅ Паттерн файлов
pattern = r"^История-затрат-Не определено-.*\.xlsx$"

# ✅ Regex для извлечения START и END из имени
# История-затрат-Не определено-{START}-{END}.xlsx
name_re = re.compile(
    r"^История-затрат-Не определено-"
    r"(?P<start>\d{4}-\d{2}-\d{2}T\d{2}_\d{2}_\d{2}\+\d{2}_\d{2})-"
    r"(?P<end>\d{4}-\d{2}-\d{2}T\d{2}_\d{2}_\d{2}\+\d{2}_\d{2})"
    r"\.xlsx$"
)

def ts_to_dt(ts: str) -> datetime:
    # '2025-11-10T00_00_00+03_00' -> '2025-11-10T00:00:00+03:00'
    iso = ts.replace("_", ":")
    iso = iso.replace("+", "+")  # оставлено для наглядности
    # после replace("_", ":") получим: '2025-11-10T00:00:00+03:00'  (именно то, что нужно)
    return datetime.fromisoformat(iso)

def pick_latest_two_files(folder: str, filename_pattern: str) -> list[str]:
    rx = re.compile(filename_pattern)
    candidates = []
    for fn in os.listdir(folder):
        if rx.match(fn):
            full = os.path.join(folder, fn)
            if os.path.isfile(full):
                candidates.append(full)

    if len(candidates) < 2:
        raise FileNotFoundError(
            f"Нужно минимум 2 файла по паттерну {filename_pattern}. Найдено: {len(candidates)}"
        )

    # Берём 2 самых новых по времени изменения
    candidates.sort(key=os.path.getmtime, reverse=False)
    return candidates[:2]

def parse_start_end(filename: str) -> tuple[str, str]:
    base = os.path.basename(filename)
    m = name_re.match(base)
    if not m:
        raise ValueError(f"Имя файла не соответствует формату: {base}")
    return m.group("start"), m.group("end")

def create_blank_excel(filename: str, sheet_title: str = "История затрат"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    wb.save(filename)

def save_df_to_named_sheet(excel_path: str, df: pd.DataFrame, sheet_name: str):
    wb = load_workbook(excel_path)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # (необязательно) удалить дефолтный лист, если он остался пустым "Sheet"
    if "Sheet" in wb.sheetnames and len(wb["Sheet"]["A"]) == 0:
        del wb["Sheet"]

    wb.save(excel_path)


# --- ОСНОВНОЙ СЦЕНАРИЙ ---

latest_two = pick_latest_two_files(downloads_folder, pattern)

# Парсим даты, чтобы правильно определить: какой файл "раньше", какой "позже"
meta = []
for fp in latest_two:
    start_ts, end_ts = parse_start_end(fp)
    meta.append((fp, start_ts, end_ts, ts_to_dt(start_ts), ts_to_dt(end_ts)))

# сортируем по start_dt: старый -> новый
meta.sort(key=lambda x: x[3])

file1_path, file1_start_ts, file1_end_ts, _, _ = meta[0]
file2_path, file2_start_ts, file2_end_ts, _, _ = meta[1]
# Читаем Excel
file1 = pd.read_excel(file1_path, engine="calamine")
file2 = pd.read_excel(file2_path, engine="calamine")

# Итоговое имя: start из file1 + end из file2
out_name = f"История-затрат-Не определено-{file1_start_ts}-{file2_end_ts}.xlsx"
excel_path = os.path.join(current_folder, out_name)

# Пишем
create_blank_excel(excel_path, sheet_title="История затрат")
save_df_to_named_sheet(
    excel_path=excel_path,
    df=pd.concat([file1, file2], ignore_index=True),
    sheet_name="История затрат",
)

print("OK:")
print("file1 =", os.path.basename(file1_path))
print("file2 =", os.path.basename(file2_path))
print("output=", excel_path)